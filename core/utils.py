import csv
import os
from os.path import join
from typing import List

import cv2
import matplotlib.colors as colors
import numpy as np
import openpyxl
from matplotlib import pyplot as plt
from matplotlib.ticker import FormatStrFormatter
from tabulate import tabulate

from core.popos import TraitsResult, LuminosityResult


# Function of rgb to hex color space
def rgb_to_hex(rgb_tuple):
    return colors.rgb2hex([1.0*x/255 for x in rgb_tuple])

# Function for generating the histogram using centered labels
def centroid_histogram(clt):
    # grab the number of different clusters and create a histogram 
    # based on the number of pixels assigned to each cluster
    numLabels = np.arange(0, len(np.unique(clt.labels_)) + 1)
    (hist, _) = np.histogram(clt.labels_, bins = numLabels)

    # normalize the histogram, such that it sums to one
    hist = hist.astype("float")
    hist /= hist.sum()

    # return the histogram
    return hist
    
# Function for plotting the histogram using centered labels    
def plot_centroid_histogram(path, clt):
    # grab the number of different clusters and create a histogram
    # based on the number of pixels assigned to each cluster
    numLabels = np.arange(0, len(np.unique(clt.labels_)) + 1)
    #(hist, _) = np.histogram(clt.labels_, bins = numLabels)

    #prepare a new figure
    fig, ax = plt.subplots()
    
    #setting the titles and lables, style
    plt.title("Dominant Color Distribution Histogram")
    plt.xlabel("Color Value")
    plt.ylabel("Percentage")
    counts, bins, patches = ax.hist(clt.labels_, numLabels, normed=1, histtype='bar', stacked=True, color='gray',alpha=0.8,edgecolor='gray')

    # Set the ticks to be at the edges of the bins.
    ax.set_xticks(bins)
    # Set the xaxis's tick labels to be formatted with 1 decimal place...
    ax.xaxis.set_major_formatter(FormatStrFormatter('%0.1f'))
    
    #centroids = sorted(clt.cluster_centers_, reverse = True)
    
    centroids = clt.cluster_centers_

    # Setting each cluster's value
    for bin_size, cbin, patch, color in zip(counts, bins, patches, centroids):
        bin_color = color.astype("uint8").tolist()
        bin_color_val = rgb_to_hex(bin_color)
        patch.set_facecolor(bin_color_val)
        patch.set_label(bin_color)
        #print (bin_color_val)
        #print len(centroids)
        #print len(counts)
    
    # add legend to a plot
    plt.legend()

        
    # Label the raw counts and the percentages below the x-axis...
    bin_centers = 0.5 * np.diff(bins) + bins[:-1]
    for count, x in zip(counts, bin_centers):
        
        # Label the percentages
        percent = '%0.000f%%' % (100 * float(count) / counts.sum())
        ax.annotate(percent, xy=(x, 0), xycoords=('data', 'axes fraction'),xytext=(0, -32), textcoords='offset points', va='top', ha='center')
        
    # Give more room at the bottom of the plot
    plt.subplots_adjust(bottom=0.15)
    
    #set the clutser distributation figure name
    #define result path for labeled images
    fig_name = path + 'color_distribution.png'
    #fig_name = 'Color_Distribution.png'
    
    #save plot as an image
    plt.savefig(fig_name)
    
    plt.close(fig)
    
        
    '''
    # Save the cluster results into text file
    file_name = 'color_data.txt'
    
    # save each clustre result as one line with space 
    counts_sorted = sorted(counts, reverse = True)
    with open(file_name, 'a') as f:
        #for idx in min(range(len(centroids)),range(len(counts))):
        for idx in range(len(counts_sorted)):
            #bin_color = color.astype("uint8").tolist()
            #bin_color_val = rgb_to_hex(bin_color)
            bin_color_val = 'Level'
            percent = '%0.000f' % (100 * float(counts_sorted[idx]) / counts.sum())
            #f.write('%s\n' % ((bin_color_val) + '   ' + str(percent)))
            f.write('%s\t' % (str(percent)))
        f.write('\n')
    '''
    
    '''
    fig = plt.figure(5)
    #clustered = np.hstack([image, quant])
    clustered = quant
    plt.imshow(clustered)
    #plt.show()

    
    #save clustered image
    filename = args["image"]
    fig_name = (str(filename[0:-4]) + '_' +'cluster_out.png')
    fig_path_save = save_path + fig_name
    mpimg.imsave(fig_path_save, clustered)
    '''
    #plt.show()
            
    return numLabels
    
# Function for plotting the color distributation as a bar figure 
def plot_colors(hist, centroids):
    # initialize the bar chart representing the relative frequency
    # of each of the colors
    bar = np.zeros((10, 100, 3), dtype = "uint8")
    startX = 0

    
    # loop over the percentage of each cluster and the color of
    # each cluster
    for (percent, color) in zip(hist, centroids):
        # plot the relative percentage of each cluster
        endX = startX + (percent * 100)
        cv2.rectangle(bar, (int(startX), 0), (int(endX), 10), color.astype("uint8").tolist(), -1)
        startX = endX
    
    # return the bar chart
    return bar

# Function for plotting the histogram using centered labels    
def plot_labeled_histogram(pixels,bins_num):
    # grab the number of different clusters and create a histogram
    # based on the number of pixels assigned to each cluster
    numLabels = bins_num
    
    #prepare a new figure
    fig, ax = plt.subplots()
    plt.title("Color Distributation Histogram")
    plt.xlabel("Color Value")
    plt.ylabel("Frequency")
    counts, bins, patches = ax.hist(pixels, numLabels, normed=1, histtype='bar', stacked=True, color='gray',alpha=0.8,edgecolor='gray')

    # Set the ticks to be at the edges of the bins.
    ax.set_xticks(bins)
    # Set the xaxis's tick labels to be formatted with 1 decimal place...
    ax.xaxis.set_major_formatter(FormatStrFormatter('%0.1f'))
    
    
    color_lable = clt.cluster_centers_

    # Setting each cluster's value
    for bin_size, cbin, patch, color in zip(counts, bins, patches, color_lable):
        bin_color = color.astype("uint8").tolist()
        bin_color_val = rgb_to_hex(bin_color)
        patch.set_facecolor(bin_color_val)
        patch.set_label(bin_color)
        #print bin_color_val
    
    # add legend to a plot
    plt.legend()
        
    # Label the raw counts and the percentages below the x-axis...
    bin_centers = 0.5 * np.diff(bins) + bins[:-1]
    for count, x in zip(counts, bin_centers):
        # Label the raw counts
        #ax.annotate(str(count), xy=(x, 0), xycoords=('data', 'axes fraction'), xytext=(0, -18), textcoords='offset points', va='top', ha='center')
        
        # Label the percentages
        percent = '%0.000f%%' % (100 * float(count) / counts.sum())
        ax.annotate(percent, xy=(x, 0), xycoords=('data', 'axes fraction'),xytext=(0, -32), textcoords='offset points', va='top', ha='center')
        
    # Give ourselves some more room at the bottom of the plot
    plt.subplots_adjust(bottom=0.15)
    
    # save plot as an image
    #plt.savefig("Color_Distribution.png")

    #plt.show()
            
    return numLabels
    

def plot_color_bar(path, bar):
    
    #show color bart and save bar figure
    fig = plt.figure(0)
    plt.title("Color Distributation Histogram")
    plt.imshow(bar)
    plt.xlabel("Percentage")
    plt.ylabel("Color category")
    frame = plt.gca()
    #frame.axes.get_xaxis().set_visible(False)
    frame.axes.get_yaxis().set_visible(False)

    #save bar image
    complete_path = path + 'color_bar.png'       
    plt.savefig(complete_path)
    plt.close(fig)


def write_luminosity_results_to_csv(results: List[LuminosityResult], output_directory: str):
    result_file = join(output_directory, 'luminous_detection.csv')

    with open(result_file, 'a+') as file:
        writer = csv.writer(file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        keys = list(results[0].keys())

        if os.stat(result_file).st_size == 0:
            writer.writerow(keys)

        for result in results:
            writer.writerow(list(result.values()))


def write_luminosity_results_to_excel(results: List[LuminosityResult], output_directory: str):
    result_file = join(output_directory, 'luminous_detection.xlsx')

    if os.path.isfile(result_file):
        wb = openpyxl.load_workbook(result_file)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        keys = list(results[0].keys())

        for i in range(3):
            sheet.cell(row=1, column=i + 1).value = keys[i]

    for result in results:
        sheet.append(list(result.values()))

    wb.save(result_file)


def write_traits_results_to_csv(results: List[TraitsResult], output_directory: str):
    result_file = join(output_directory, 'traits.csv')

    with open(result_file, 'a+') as file:
        writer = csv.writer(file, delimiter=',', quotechar='|', quoting=csv.QUOTE_MINIMAL)
        keys = list(results[0].keys())

        if os.stat(result_file).st_size == 0:
            writer.writerow(keys)

        for result in results:
            writer.writerow(list(result.values()))


def write_traits_results_to_excel(results: List[TraitsResult], output_directory: str):
    result_file = join(output_directory, 'traits.xlsx')

    if os.path.isfile(result_file):
        wb = openpyxl.load_workbook(result_file)
        sheet = wb.active
    else:
        wb = openpyxl.Workbook()
        sheet = wb.active
        keys = list(results[0].keys())

        for i in range(7):
            sheet.cell(row=1, column=i + 1).value = keys[i]

    for result in results:
        sheet.append(list(result.values()))

    wb.save(result_file)


def outlier_double_mad(data, thresh=3.5):
    """
    FOR ASYMMETRIC DISTRIBUTION
    Returns : filtered array excluding the outliers

    Parameters : the actual data Points array

    Calculates median to divide data into 2 halves.(skew conditions handled)
    Then those two halves are treated as separate data with calculation same as for symmetric distribution.(first answer)
    Only difference being , the thresholds are now the median distance of the right and left median with the actual data median

    Warning: this function does not check for NAs,
    nor does it address issues when more than 50% of your data have identical values
    """

    m = np.median(data)
    abs_dev = np.abs(data - m)
    left_mad = np.median(abs_dev[data <= m])
    right_mad = np.median(abs_dev[data >= m])
    data_mad = left_mad * np.ones(len(data))
    data_mad[data > m] = right_mad
    modified_z_score = 0.6745 * abs_dev / data_mad
    modified_z_score[data == m] = 0
    return modified_z_score > thresh


def get_cmap(n, name='hsv'):
    """
    Returns a function that maps each index in 0, 1, ..., n-1 to a distinct
    RGB color; the keyword argument name must be a standard mpl colormap name.
    """
    return plt.cm.get_cmap(name, n)


def rgb2hex(color):
    return "#{:02x}{:02x}{:02x}".format(int(color[0]), int(color[1]), int(color[2]))