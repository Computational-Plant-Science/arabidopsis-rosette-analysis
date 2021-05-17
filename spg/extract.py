import argparse
import csv
import glob
import os
import traceback
import warnings
from collections import Counter
from contextlib import closing
from multiprocessing import Pool
from os.path import join
from typing import List

import cv2
import imutils
import matplotlib.pyplot as plt
import numpy as np
import openpyxl
from scipy import ndimage
from scipy.interpolate import interp1d
from scipy.spatial import distance as dist
from skan import Skeleton, summarize, draw
from skimage import img_as_float, img_as_ubyte, img_as_bool
from skimage import morphology
from skimage.color import rgb2lab, deltaE_cie76
from skimage.feature import peak_local_max
from skimage.segmentation import watershed
from tabulate import tabulate

from spg.color_seg import color_cluster_seg
from spg.curvature import ComputeCurvature
from spg.utils import outlier_double_mad, get_cmap, rgb2hex
from spg.popos import SPGOptions, TraitsResult

warnings.filterwarnings("ignore")

MBFACTOR = float(1 << 20)


def skeleton_bw(thresh):
    # Convert mask to boolean image, rather than 0 and 255 for skimage to use it

    # convert an image from OpenCV to skimage
    thresh_sk = img_as_float(thresh)

    image_bw = img_as_bool((thresh_sk))

    skeleton = morphology.skeletonize(image_bw)

    skeleton_img = skeleton.astype(np.uint8) * 255

    return skeleton_img, skeleton


def watershed_seg(thresh, min_distance_value):
    # compute the exact Euclidean distance from every binary
    # pixel to the nearest zero pixel, then find peaks in this distance map
    D = ndimage.distance_transform_edt(thresh)

    localMax = peak_local_max(D, indices=False, min_distance=min_distance_value, labels=thresh)

    # perform a connected component analysis on the local peaks,
    # using 8-connectivity, then apply the Watershed algorithm
    markers = ndimage.label(localMax, structure=np.ones((3, 3)))[0]

    labels = watershed(-D, markers, mask=thresh)

    print(f"{len(np.unique(labels)) - 1} unique segments found")
    return labels


def individual_object_seg(orig, labels, save_path, base_name, file_extension, leaf_images: bool = True):
    (width, height, n_channel) = orig.shape

    for label in np.unique(labels):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue

        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros((width, height), dtype="uint8")
        mask[labels == label] = 255

        # apply individual object mask
        masked = cv2.bitwise_and(orig, orig, mask=mask)

        if leaf_images:
            result_img_path = (save_path + base_name + '_leaf_' + str(label) + file_extension)
            cv2.imwrite(result_img_path, masked)


def comp_external_contour(orig, thresh):
    # find contours and get the external one
    contours, hier = cv2.findContours(thresh, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    img_height, img_width, img_channels = orig.shape

    index = 1

    for c in contours:

        # get the bounding rect
        x, y, w, h = cv2.boundingRect(c)

        if w > img_width * 0.01 and h > img_height * 0.01:

            trait_img = cv2.drawContours(orig, contours, -1, (255, 255, 0), 1)

            # draw a green rectangle to visualize the bounding rect
            roi = orig[y:y + h, x:x + w]

            print("ROI {} detected ...\n".format(index))
            # result_file = (save_path +  str(index) + file_extension)
            # cv2.imwrite(result_file, roi)

            trait_img = cv2.rectangle(orig, (x, y), (x + w, y + h), (255, 255, 0), 3)

            index += 1

            '''
            #get the min area rect
            rect = cv2.minAreaRect(c)
            box = cv2.boxPoints(rect)
            # convert all coordinates floating point values to int
            box = np.int0(box)
            #draw a red 'nghien' rectangle
            trait_img = cv2.drawContours(orig, [box], 0, (0, 0, 255))
            '''
            # get convex hull
            hull = cv2.convexHull(c)
            # draw it in red color
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255), 3)

            '''
            # calculate epsilon base on contour's perimeter
            # contour's perimeter is returned by cv2.arcLength
            epsilon = 0.01 * cv2.arcLength(c, True)
            # get approx polygons
            approx = cv2.approxPolyDP(c, epsilon, True)
            # draw approx polygons
            trait_img = cv2.drawContours(orig, [approx], -1, (0, 255, 0), 1)
         
            # hull is convex shape as a polygon
            hull = cv2.convexHull(c)
            trait_img = cv2.drawContours(orig, [hull], -1, (0, 0, 255))
            '''

            '''
            #get the min enclosing circle
            (x, y), radius = cv2.minEnclosingCircle(c)
            # convert all values to int
            center = (int(x), int(y))
            radius = int(radius)
            # and draw the circle in blue
            trait_img = cv2.circle(orig, center, radius, (255, 0, 0), 2)
            '''

            area = cv2.contourArea(c)
            print("Leaf area = {0:.2f}... \n".format(area))

            hull = cv2.convexHull(c)
            hull_area = cv2.contourArea(hull)
            solidity = float(area) / hull_area
            print("solidity = {0:.2f}... \n".format(solidity))

            extLeft = tuple(c[c[:, :, 0].argmin()][0])
            extRight = tuple(c[c[:, :, 0].argmax()][0])
            extTop = tuple(c[c[:, :, 1].argmin()][0])
            extBot = tuple(c[c[:, :, 1].argmax()][0])

            trait_img = cv2.circle(orig, extLeft, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extRight, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extTop, 3, (255, 0, 0), -1)
            trait_img = cv2.circle(orig, extBot, 3, (255, 0, 0), -1)

            max_width = dist.euclidean(extLeft, extRight)
            max_height = dist.euclidean(extTop, extBot)

            if max_width > max_height:
                trait_img = cv2.line(orig, extLeft, extRight, (0, 255, 0), 2)
            else:
                trait_img = cv2.line(orig, extTop, extBot, (0, 255, 0), 2)

            print("Width and height are {0:.2f},{1:.2f}... \n".format(w, h))

    return trait_img, area, solidity, w, h


def compute_curv(orig, labels):
    gray = cv2.cvtColor(orig, cv2.COLOR_BGR2GRAY)

    curv_sum = 0.0
    count = 0
    # curvature computation
    # loop over the unique labels returned by the Watershed algorithm
    for index, label in enumerate(np.unique(labels), start=1):
        # if the label is zero, we are examining the 'background'
        # so simply ignore it
        if label == 0:
            continue

        # otherwise, allocate memory for the label region and draw
        # it on the mask
        mask = np.zeros(gray.shape, dtype="uint8")
        mask[labels == label] = 255

        # detect contours in the mask and grab the largest one
        # cnts = cv2.findContours(mask.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)[-2]
        contours, hierarchy = cv2.findContours(mask.copy(), cv2.RETR_TREE, cv2.CHAIN_APPROX_SIMPLE)
        c = max(contours, key=cv2.contourArea)

        # draw a circle enclosing the object
        ((x, y), r) = cv2.minEnclosingCircle(c)
        label_trait = cv2.circle(orig, (int(x), int(y)), 3, (0, 255, 0), 2)
        label_trait = cv2.putText(orig, "#{}".format(label), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)
        # cv2.putText(orig, "#{}".format(curvature), (int(x) - 10, int(y)), cv2.FONT_HERSHEY_SIMPLEX, 0.6, (0, 0, 255), 2)

        if len(c) >= 5:
            try:
                label_trait = cv2.drawContours(orig, [c], -1, (255, 0, 0), 2)
                ellipse = cv2.fitEllipse(c)
                label_trait = cv2.ellipse(orig, ellipse, (0, 255, 0), 2)

                c_np = np.vstack(c).squeeze()
                count += 1

                x = c_np[:, 0]
                y = c_np[:, 1]

                comp_curv = ComputeCurvature(x, y)
                curvature = comp_curv.fit(x, y)

                curv_sum = curv_sum + curvature
            except:
                print(traceback.format_exc())
        else:
            # optional to "delete" the small contours
            label_trait = cv2.drawContours(orig, [c], -1, (0, 0, 255), 2)
            print("lack of enough points to fit ellipse")

    if count > 0:
        print('average curvature = {0:.2f}\n'.format(curv_sum / count))
    else:
        count = 1.0

    return curv_sum / count, label_trait


def color_region(image, mask, output_directory, file_name, num_clusters):
    # read the image
    # grab image width and height
    (h, w) = image.shape[:2]

    # apply the mask to get the segmentation of plant
    masked_image_ori = cv2.bitwise_and(image, image, mask=mask)

    # define result path for labeled images
    result_img_path = join(output_directory, f"{file_name}.masked.png")
    cv2.imwrite(result_img_path, masked_image_ori)

    # convert to RGB
    image_RGB = cv2.cvtColor(masked_image_ori, cv2.COLOR_BGR2RGB)

    # reshape the image to a 2D array of pixels and 3 color values (RGB)
    pixel_values = image_RGB.reshape((-1, 3))

    # convert to float
    pixel_values = np.float32(pixel_values)

    # define stopping criteria
    criteria = (cv2.TERM_CRITERIA_EPS + cv2.TERM_CRITERIA_MAX_ITER, 100, 0.2)

    # number of clusters (K)
    # num_clusters = 5
    compactness, labels, (centers) = cv2.kmeans(pixel_values, num_clusters, None, criteria, 10, cv2.KMEANS_RANDOM_CENTERS)

    # convert back to 8 bit values
    centers = np.uint8(centers)

    # flatten the labels array
    labels_flat = labels.flatten()

    # convert all pixels to the color of the centroids
    segmented_image = centers[labels_flat]

    # reshape back to the original image dimension
    segmented_image = segmented_image.reshape(image_RGB.shape)

    segmented_image_BRG = cv2.cvtColor(segmented_image, cv2.COLOR_RGB2BGR)
    # define result path for labeled images
    result_img_path = join(output_directory, f"{file_name}.clustered.png")
    cv2.imwrite(result_img_path, segmented_image_BRG)

    '''
    fig = plt.figure()
    ax = Axes3D(fig)        
    for label, pix in zip(labels, segmented_image):
        ax.scatter(pix[0], pix[1], pix[2], color = (centers))
            
    result_file = (save_path + base_name + 'color_cluster_distributation.png')
    plt.savefig(result_file)
    '''
    # Show only one chosen cluster
    # masked_image = np.copy(image)
    masked_image = np.zeros_like(image_RGB)

    # convert to the shape of a vector of pixel values
    masked_image = masked_image.reshape((-1, 3))
    # color (i.e cluster) to render
    # cluster = 2

    cmap = get_cmap(num_clusters + 1)

    # clrs = sns.color_palette('husl', n_colors = num_clusters)  # a list of RGB tuples

    color_conversion = interp1d([0, 1], [0, 255])

    for cluster in range(num_clusters):

        print("Processing Cluster{0} ...\n".format(cluster))
        # print(clrs[cluster])
        # print(color_conversion(clrs[cluster]))

        masked_image[labels_flat == cluster] = centers[cluster]

        # print(centers[cluster])

        # convert back to original shape
        masked_image_rp = masked_image.reshape(image_RGB.shape)

        # masked_image_BRG = cv2.cvtColor(masked_image, cv2.COLOR_RGB2BGR)
        # cv2.imwrite('maksed.png', masked_image_BRG)

        gray = cv2.cvtColor(masked_image_rp, cv2.COLOR_BGR2GRAY)

        # threshold the image, then perform a series of erosions +
        # dilations to remove any small regions of noise
        thresh = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY)[1]

        # thresh_cleaned = clear_border(thresh)

        cnts = cv2.findContours(thresh.copy(), cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        cnts = imutils.grab_contours(cnts)
        # c = max(cnts, key=cv2.contourArea)

        '''
        # compute the center of the contour area and draw a circle representing the center
        M = cv2.moments(c)
        cX = int(M["m10"] / M["m00"])
        cY = int(M["m01"] / M["m00"])
        # draw the countour number on the image
        result = cv2.putText(masked_image_rp, "#{}".format(cluster + 1), (cX - 20, cY), cv2.FONT_HERSHEY_SIMPLEX, 1.0, (255, 255, 255), 2)
        '''

        if not cnts:
            print("findContours is empty")
        else:

            # loop over the (unsorted) contours and draw them
            for (i, c) in enumerate(cnts):
                result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(np.random.random(3)), 2)
                # result = cv2.drawContours(masked_image_rp, c, -1, color_conversion(clrs[cluster]), 2)

            # result = result(np.where(result == 0)== 255)
            result[result == 0] = 255

            result_BRG = cv2.cvtColor(result, cv2.COLOR_RGB2BGR)
            result_img_path = join(output_directory, f"{file_name}.result.{cluster}.png")
            cv2.imwrite(result_img_path, result_BRG)

    counts = Counter(labels_flat)
    # sort to ensure correct color percentage
    counts = dict(sorted(counts.items()))

    center_colors = centers

    # We get ordered colors by iterating through the keys
    ordered_colors = [center_colors[i] for i in counts.keys()]
    hex_colors = [rgb2hex(ordered_colors[i]) for i in counts.keys()]
    rgb_colors = [ordered_colors[i] for i in counts.keys()]

    # print(hex_colors)

    index_bkg = [index for index in range(len(hex_colors)) if hex_colors[index] == '#000000']

    # print(index_bkg[0])

    # print(counts)
    # remove background color
    del hex_colors[index_bkg[0]]
    del rgb_colors[index_bkg[0]]

    # Using dictionary comprehension to find list 
    # keys having value . 
    delete = [key for key in counts if key == index_bkg[0]]

    # delete the key 
    for key in delete: del counts[key]

    fig = plt.figure(figsize=(6, 6))
    plt.pie(counts.values(), labels=hex_colors, colors=hex_colors)

    # define result path for labeled images
    result_img_path = join(output_directory, f"{file_name}.pie_color.png")
    plt.savefig(result_img_path)

    return rgb_colors


def trait_extract(options: SPGOptions) -> List[TraitsResult]:
    if options.multiprocessing:
        cpus = os.cpu_count()
        print(f"Using up to {cpus} processes to extract traits from {len(options.input_images)} images")
        with closing(Pool(processes=cpus)) as pool:
            traits_results = pool.map(trait_extract, inputs)
            pool.terminate()
    else:
        print(f"Using a single process to extract traits from {len(files)} images")
        traits_results = [trait_extract(image) for image in inputs]

    try:
        _, file_extension = os.path.splitext(options.input_file)
        file_size = os.path.getsize(options.input_file) / MBFACTOR

        print("Segmenting plant object using automatic color clustering method")
        if (file_size > 5.0):
            print(f"It may take some time due to large file size ({file_size} MB)")

        args_colorspace = 'lab'
        args_channels = '1'
        args_num_clusters = 2

        image = cv2.imread(options.input_file)

        # circle detection
        # _, circles, cropped = circle_detect(options)
        image_copy = image.copy()

        # color clustering based plant object segmentation
        segmented = color_cluster_seg(image_copy, args_colorspace, args_channels, args_num_clusters)
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_seg{file_extension}"), segmented)

        num_clusters = 5
        # save color quantization result
        # rgb_colors = color_quantization(image, thresh, save_path, num_clusters)
        rgb_colors = color_region(image_copy, segmented, options.output_directory + '/', options.input_stem, num_clusters)

        selected_color = rgb2lab(np.uint8(np.asarray([[rgb_colors[0]]])))

        print("Color difference are : ")

        print(selected_color)

        color_diff = []

        for index, value in enumerate(rgb_colors):
            # print(index, value)
            curr_color = rgb2lab(np.uint8(np.asarray([[value]])))
            diff = deltaE_cie76(selected_color, curr_color)

            color_diff.append(diff)

            print(index, value, diff)

            ###############################################

        # accquire medial axis of segmentation mask
        # image_medial_axis = medial_axis_image(thresh)

        image_skeleton, skeleton = skeleton_bw(segmented)

        # save _skeleton result
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_skeleton{file_extension}"), img_as_ubyte(image_skeleton))

        ###
        # ['skeleton-id', 'node-id-src', 'node-id-dst', 'branch-distance',
        # 'branch-type', 'mean-pixel-value', 'stdev-pixel-value',
        # 'image-coord-src-0', 'image-coord-src-1', 'image-coord-dst-0', 'image-coord-dst-1',
        # 'coord-src-0', 'coord-src-1', 'coord-dst-0', 'coord-dst-1', 'euclidean-distance']
        ###

        # get brach data
        branch_data = summarize(Skeleton(image_skeleton))
        # print(branch_data)

        # select end branch
        sub_branch = branch_data.loc[branch_data['branch-type'] == 1]

        sub_branch_branch_distance = sub_branch["branch-distance"].tolist()

        # remove outliers in branch distance
        outlier_list = outlier_double_mad(sub_branch_branch_distance, thresh=3.5)

        indices = [i for i, x in enumerate(outlier_list) if x]

        sub_branch_cleaned = sub_branch.drop(sub_branch.index[indices])

        # print(outlier_list)
        # print(indices)
        # print(sub_branch)

        print(sub_branch_cleaned)

        '''
        min_distance_value_list = sub_branch_cleaned["branch-distance"].tolist()

        min_distance_value_list.sort()

        min_distance_value = int(min_distance_value_list[2])

        print("Smallest branch-distance is:", min_distance_value)

        #fig = plt.plot()

        (img_endpt_overlay, img_marker) = overlay_skeleton_endpoints(source_image, sub_branch_cleaned)

        result_file = (save_path + base_name + '_endpts_overlay' + file_extension)
        plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
        plt.close()

        result_file = (save_path + base_name + '_marker' + file_extension)
        cv2.imwrite(result_file, img_marker)
        '''

        branch_type_list = sub_branch_cleaned["branch-type"].tolist()

        # print(branch_type_list.count(1))

        print("[INFO] {} branch end points found\n".format(branch_type_list.count(1)))

        # img_hist = branch_data.hist(column = 'branch-distance', by = 'branch-type', bins = 100)
        # result_file = (save_path + base_name + '_hist' + file_extension)
        # plt.savefig(result_file, transparent = True, bbox_inches = 'tight', pad_inches = 0)
        # plt.close()

        fig = plt.plot()
        source_image = cv2.cvtColor(image_copy, cv2.COLOR_BGR2RGB)
        # img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source = 'branch-distance', skeleton_colormap = 'hsv')
        img_overlay = draw.overlay_euclidean_skeleton_2d(source_image, branch_data, skeleton_color_source='branch-type', skeleton_colormap='hsv')
        plt.savefig(join(options.output_directory, f"{options.input_stem}_euclidean_graph_overlay{file_extension}"), transparent=True,
                    bbox_inches='tight', pad_inches=0)
        plt.close()

        ############################################## leaf number computation
        min_distance_value = 20
        # watershed based leaf area segmentaiton
        labels = watershed_seg(segmented, min_distance_value)

        # labels = watershed_seg_marker(orig, thresh, min_distance_value, img_marker)

        individual_object_seg(image_copy, labels, options.output_directory + '/', options.input_stem, file_extension)

        # save watershed result label image
        # Map component labels to hue val
        label_hue = np.uint8(128 * labels / np.max(labels))
        # label_hue[labels == largest_label] = np.uint8(15)
        blank_ch = 255 * np.ones_like(label_hue)
        labeled_img = cv2.merge([label_hue, blank_ch, blank_ch])

        # cvt to BGR for display
        labeled_img = cv2.cvtColor(labeled_img, cv2.COLOR_HSV2BGR)

        # set background label to black
        labeled_img[label_hue == 0] = 0
        # plt.imsave(result_file, img_as_float(labels), cmap = "Spectral")
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_label{file_extension}"), labeled_img)

        (avg_curv, label_trait) = compute_curv(image_copy, labels)

        # save watershed result label image
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_curv{file_extension}"), label_trait)

        # find external contour
        (trait_img, area, solidity, max_width, max_height) = comp_external_contour(image_copy, segmented)
        # save segmentation result
        # print(filename)
        cv2.imwrite(join(options.output_directory, f"{options.input_stem}_excontour{file_extension}"), trait_img)

        n_leaves = int(len(np.unique(labels)) / 1 - 1)

        # print("[INFO] {} n_leaves found\n".format(len(np.unique(labels)) - 1))

        # Path("/tmp/d/a.dat").name

        return TraitsResult(options.input_stem, False, area, solidity, max_width, max_height, avg_curv, n_leaves)
    except:
        print(f"Error in trait extraction: {traceback.format_exc()}")
        return TraitsResult(options.input_stem, True, None, None, None, None, None, None)


if __name__ == '__main__':

    ap = argparse.ArgumentParser()
    ap.add_argument("-p", "--path", required=True, help="path to image file")
    ap.add_argument("-ft", "--filetype", required=True, help="Image filetype")
    ap.add_argument("-r", "--result", required=False, help="result path")
    ap.add_argument('-s', '--color-space', type=str, default='lab', help='Color space to use: BGR (default), HSV, Lab, YCrCb (YCC)')
    ap.add_argument('-c', '--channels', type=str, default='1', help='Channel indices to use for clustering, where 0 is the first channel,'
                                                                    + ' 1 is the second channel, etc. E.g., if BGR color space is used, "02" '
                                                                    + 'selects channels B and R. (default "all")')
    ap.add_argument('-n', '--num-clusters', type=int, default=2, help='Number of clusters for K-means clustering (default 2, min 2).')
    args = vars(ap.parse_args())

    # setting path to model file
    file_path = args["path"]
    ext = args['filetype']

    # accquire image file list
    filetype = '*.' + ext
    image_file_path = file_path + filetype

    # accquire image file list
    imgList = sorted(glob.glob(image_file_path))

    # print((imgList))
    # global save_path

    n_images = len(imgList)

    result_list = []

    # loop execute
    for image in imgList:
        (filename, area, solidity, max_width, max_height, avg_curv, n_leaves) = extract_traits(image)

        result_list.append([filename, area, solidity, max_width, max_height, avg_curv, n_leaves])

    # print(result_list)

    '''
    # get cpu number for parallel processing
    agents = psutil.cpu_count()   
    #agents = multiprocessing.cpu_count() 
    #agents = 8
    
    print("Using {0} cores to perfrom parallel processing... \n".format(int(agents)))
    
    # Create a pool of processes. By default, one is created for each CPU in the machine.
    # extract the bouding box for each image in file list
    with closing(Pool(processes = agents)) as pool:
        result_list = pool.map(extract_traits, imgList)
        pool.terminate()
    '''

    # trait_file = (os.path.dirname(os.path.abspath(file_path)) + '/' + 'trait.xlsx')

    print("Summary: {0} plant images were processed...\n".format(n_images))

    # output in command window in a sum table

    table = tabulate(result_list, headers=['filename', 'area', 'solidity', 'max_width', 'max_height', 'avg_curv', 'n_leaves'], tablefmt='orgtbl')

    print(table + "\n")

    if (args['result']):

        trait_file = (args['result'] + 'trait.xlsx')
        trait_file_csv = (args['result'] + 'trait.csv')
    else:
        trait_file = (file_path + 'trait.xlsx')
        trait_file_csv = (file_path + 'trait.csv')

    if os.path.isfile(trait_file):
        # update values
        # Open an xlsx for reading
        wb = openpyxl.load_workbook(trait_file)

        # Get the current Active Sheet
        sheet = wb.active

    else:
        # Keep presets
        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet.cell(row=1, column=1).value = 'filename'
        sheet.cell(row=1, column=2).value = 'leaf_area'
        sheet.cell(row=1, column=3).value = 'solidity'
        sheet.cell(row=1, column=4).value = 'max_width'
        sheet.cell(row=1, column=5).value = 'max_height'
        sheet.cell(row=1, column=6).value = 'curvature'
        sheet.cell(row=1, column=7).value = 'number_leaf'

    for row in result_list:
        sheet.append(row)

    # save the csv file
    wb.save(trait_file)

    wb = openpyxl.load_workbook(trait_file)
    sh = wb.active  # was .get_active_sheet()
    with open(trait_file_csv, 'w', newline="") as f:
        c = csv.writer(f)
        for r in sh.rows:  # generator; was sh.rows
            c.writerow([cell.value for cell in r])
